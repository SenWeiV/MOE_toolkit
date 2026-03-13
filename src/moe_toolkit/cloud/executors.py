"""Execution backends for MOE cloud runs."""

from __future__ import annotations

import asyncio
import csv
import json
import shutil
from dataclasses import dataclass
from pathlib import Path
from typing import Awaitable, Callable, Protocol

from openpyxl import Workbook, load_workbook

from moe_toolkit.schemas.common import RemoteTaskRequest, RoutePlan


@dataclass(slots=True)
class ExecutionContext:
    """Execution context shared by all backends."""

    run_id: str
    request: RemoteTaskRequest
    route_plan: RoutePlan
    run_root: Path
    input_dir: Path
    artifacts_dir: Path


class ExecutionBackend(Protocol):
    """Protocol for a run execution backend."""

    async def execute(self, context: ExecutionContext) -> None:
        """Executes the run and writes artifacts into the run workspace."""


class InlineExecutor:
    """Current in-process executor used for tests and local development."""

    async def execute(self, context: ExecutionContext) -> None:
        selected_tools = context.route_plan.selected_tools or [
            image.removeprefix("moe-tool-")
            for image in context.route_plan.selected_images
        ]
        for source in sorted(context.input_dir.iterdir()):
            if source.suffix.lower() not in {".csv", ".tsv", ".xlsx"}:
                continue
            if any(tool in selected_tools for tool in ["pandas", "openpyxl"]):
                self._create_summary_artifacts(context, source)
            if "matplotlib" in selected_tools and source.suffix.lower() in {".csv", ".tsv"}:
                self._create_chart_artifact(context, source)
            if "openpyxl" in selected_tools:
                self._create_spreadsheet_artifact(context, source)
        if "markdown-report" in selected_tools:
            self._create_markdown_report(context)
        await asyncio.sleep(0)

    def _create_summary_artifacts(self, context: ExecutionContext, source: Path) -> None:
        fieldnames, rows = self._load_rows(source)

        summary = {
            "source": source.name,
            "rows": len(rows),
            "columns": fieldnames,
            "numeric_columns": self._find_numeric_columns(rows, fieldnames),
            "task": context.request.task,
        }
        summary_path = context.artifacts_dir / f"{source.stem}-summary.json"
        summary_path.write_text(
            json.dumps(summary, indent=2, ensure_ascii=False) + "\n",
            encoding="utf-8",
        )

    def _create_chart_artifact(self, context: ExecutionContext, source: Path) -> None:
        fieldnames, rows = self._load_rows(source)
        numeric_columns = self._find_numeric_columns(rows, fieldnames)
        if not numeric_columns:
            return
        chart_path = context.artifacts_dir / f"{source.stem}-chart.svg"
        chart_path.write_text(
            self._build_svg_chart(rows, numeric_columns[0]),
            encoding="utf-8",
        )

    def _create_spreadsheet_artifact(self, context: ExecutionContext, source: Path) -> None:
        fieldnames, rows = self._load_rows(source)
        workbook = Workbook()
        summary_sheet = workbook.active
        summary_sheet.title = "summary"
        summary_sheet.append(["source", source.name])
        summary_sheet.append(["task", context.request.task])
        summary_sheet.append(["rows", len(rows)])
        summary_sheet.append(["columns", ", ".join(fieldnames)])
        summary_sheet.append(["numeric_columns", ", ".join(self._find_numeric_columns(rows, fieldnames))])
        data_sheet = workbook.create_sheet("data")
        if fieldnames:
            data_sheet.append(fieldnames)
        for row in rows:
            data_sheet.append([row.get(field, "") for field in fieldnames])
        workbook.save(context.artifacts_dir / f"{source.stem}-report.xlsx")

    def _create_markdown_report(self, context: ExecutionContext) -> None:
        lines = [
            "# MOE Toolkit Run Report",
            "",
            f"- Run ID: {context.run_id}",
            f"- Task: {context.request.task}",
            "",
            "## Artifacts",
        ]
        artifact_names = sorted(path.name for path in context.artifacts_dir.iterdir() if path.is_file())
        if artifact_names:
            lines.extend(f"- {name}" for name in artifact_names)
        else:
            lines.append("- None")
        (context.artifacts_dir / "run-report.md").write_text(
            "\n".join(lines).rstrip() + "\n",
            encoding="utf-8",
        )

    @staticmethod
    def _load_rows(source: Path) -> tuple[list[str], list[dict[str, str]]]:
        if source.suffix.lower() == ".xlsx":
            workbook = load_workbook(source, read_only=True, data_only=True)
            sheet = workbook.active
            raw_rows = list(sheet.iter_rows(values_only=True))
            if not raw_rows:
                return [], []
            fieldnames = [str(value or "").strip() for value in raw_rows[0]]
            rows: list[dict[str, str]] = []
            for raw_row in raw_rows[1:]:
                row = {
                    fieldnames[index]: "" if value is None else str(value)
                    for index, value in enumerate(raw_row)
                    if index < len(fieldnames) and fieldnames[index]
                }
                if row:
                    rows.append(row)
            return fieldnames, rows
        delimiter = "\t" if source.suffix.lower() == ".tsv" else ","
        with source.open("r", encoding="utf-8", newline="") as handle:
            reader = csv.DictReader(handle, delimiter=delimiter)
            rows = list(reader)
            fieldnames = reader.fieldnames or []
        return fieldnames, rows

    @staticmethod
    def _find_numeric_columns(rows: list[dict[str, str]], fieldnames: list[str]) -> list[str]:
        numeric_columns: list[str] = []
        for field in fieldnames:
            values = [row.get(field, "").strip() for row in rows if row.get(field, "").strip()]
            if values and all(InlineExecutor._is_number(value) for value in values[:20]):
                numeric_columns.append(field)
        return numeric_columns

    @staticmethod
    def _is_number(value: str) -> bool:
        try:
            float(value)
        except ValueError:
            return False
        return True

    @staticmethod
    def _build_svg_chart(rows: list[dict[str, str]], field: str) -> str:
        values = [float(row[field]) for row in rows if row.get(field, "").strip()]
        if not values:
            values = [0.0]
        max_value = max(values) or 1.0
        points: list[str] = []
        width = 480
        height = 240
        for index, value in enumerate(values):
            x = 20 + (index * (width - 40) / max(len(values) - 1, 1))
            y = height - 20 - (value / max_value) * (height - 40)
            points.append(f"{x:.2f},{y:.2f}")
        polyline = " ".join(points)
        return "\n".join(
            [
                '<svg xmlns="http://www.w3.org/2000/svg" width="480" height="240">',
                '<rect width="100%" height="100%" fill="white" />',
                f'<text x="20" y="24" font-size="16">MOE Preview Chart: {field}</text>',
                f'<polyline fill="none" stroke="#2563eb" stroke-width="2" points="{polyline}" />',
                "</svg>",
                "",
            ]
        )


class DockerRunner(Protocol):
    """Command runner used by the docker executor."""

    async def __call__(self, command: list[str]) -> None:
        """Runs a docker command and raises on failure."""


async def run_docker_command(command: list[str]) -> None:
    """Default docker command runner."""

    process = await asyncio.create_subprocess_exec(
        *command,
        stdout=asyncio.subprocess.PIPE,
        stderr=asyncio.subprocess.PIPE,
    )
    stdout, stderr = await process.communicate()
    if process.returncode != 0:
        raise RuntimeError(
            "Docker command failed: "
            + " ".join(command)
            + f"\nstdout={stdout.decode(errors='ignore')}\nstderr={stderr.decode(errors='ignore')}"
        )


class DockerExecutor:
    """Executor that runs curated tool images inside Docker."""

    def __init__(
        self,
        docker_binary: str = "docker",
        storage_root: Path | None = None,
        host_storage_root: Path | None = None,
        network_mode: str = "none",
        runner: DockerRunner = run_docker_command,
    ) -> None:
        self.docker_binary = docker_binary
        self.storage_root = storage_root
        self.host_storage_root = host_storage_root
        self.network_mode = network_mode
        self.runner = runner

    async def execute(self, context: ExecutionContext) -> None:
        mount_source = self._resolve_mount_source(context.run_root)
        for image in context.route_plan.selected_images:
            command = [
                self.docker_binary,
                "run",
                "--rm",
                "--network",
                self.network_mode,
                "-v",
                f"{mount_source}:/work",
                "-w",
                "/work",
                image,
            ]
            await self.runner(command)

    def _resolve_mount_source(self, run_root: Path) -> Path:
        if self.storage_root is None or self.host_storage_root is None:
            return run_root
        try:
            relative_run_path = run_root.relative_to(self.storage_root)
        except ValueError:
            return run_root
        return self.host_storage_root / relative_run_path


def detect_media_type(path: Path) -> str:
    """Infers an artifact media type from the file suffix."""

    suffix = path.suffix.lower()
    if suffix == ".json":
        return "application/json"
    if suffix == ".svg":
        return "image/svg+xml"
    if suffix == ".md":
        return "text/markdown"
    if suffix == ".xlsx":
        return "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    return "application/octet-stream"


def prepare_run_workspace(
    storage_root: Path,
    run_id: str,
    request: RemoteTaskRequest,
    upload_paths: dict[str, Path],
    route_plan: RoutePlan,
) -> ExecutionContext:
    """Creates the workspace directory structure for a run."""

    run_root = storage_root / "runs" / run_id
    input_dir = run_root / "inputs"
    artifacts_dir = run_root / "artifacts"
    input_dir.mkdir(parents=True, exist_ok=True)
    artifacts_dir.mkdir(parents=True, exist_ok=True)

    attachment_entries: list[dict[str, str]] = []
    for index, upload_id in enumerate(request.attachments, start=1):
        source = upload_paths[upload_id]
        target_name = f"{index:02d}-{source.name}"
        target_path = input_dir / target_name
        shutil.copy2(source, target_path)
        attachment_entries.append(
            {
                "upload_id": upload_id,
                "source_name": source.name,
                "workspace_name": target_name,
            }
        )

    run_payload = {
        "run_id": run_id,
        "task": request.task,
        "session_id": request.session_id,
        "attachments": attachment_entries,
        "output_preferences": request.output_preferences,
        "route_plan": route_plan.model_dump(mode="json"),
    }
    (run_root / "run.json").write_text(
        json.dumps(run_payload, indent=2, ensure_ascii=False) + "\n",
        encoding="utf-8",
    )
    return ExecutionContext(
        run_id=run_id,
        request=request,
        route_plan=route_plan,
        run_root=run_root,
        input_dir=input_dir,
        artifacts_dir=artifacts_dir,
    )
